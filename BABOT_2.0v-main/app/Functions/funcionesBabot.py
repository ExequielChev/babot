# --------------------------------------------------------------------------------------------------------------
# BABOT - Asistente virtual creado en Python por EXCEL-ENTE 2023
# Desarrollador : Kevin Turkienich
# Contacto : Kevin_turkienich@outlook.com
# --------------------------------------------------------------------------------------------------------------
    # IMPORTACION DE MODULOS
# --------------------------------------------------------------------------------------------------------------
import logging
import os
from tkinter import messagebox
from openpyxl import load_workbook
import pandas as pd
# --------------------------------------------------------------------------------------------------------------

# Funciones de validacion:

def verificar_existencia_hoja(archivo_excel, nombre_hoja):
    try:
        with pd.ExcelFile(archivo_excel) as excel:
            hojas = excel.sheet_names
            if nombre_hoja in hojas:
                return True
            else:
                return False
    except Exception as e:
        logging.info(f'Error al verificar la existencia de la hoja: {e}')
        return False
    
# --------------------------------------------------------------------------------------------------------------

################################################################################################################
# Funcion de limpieza de carpeta ouput
def LIMPIAR_OUTPUT_0001(pasos_bot):

    logging.warning(' ** Ejecutando LIMPIAR_OUTPUT_0001.. **')
    logging.warning('')

    continuar = True

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de existencia de variables
    # --------------------------------------------------------------------------------------------------------------
    if continuar == True:

        ACTIVE_0001 = pasos_bot.get('ACTIVE_0001')
        if ACTIVE_0001 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave ACTIVE_0001 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave ACTIVE_0002 esta en el diccionario.')


        PATH_SALIDA_0001 = pasos_bot.get('PATH_SALIDA_0001')
        if PATH_SALIDA_0001 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave PATH_SALIDA_0001 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave PATH_ENTRADA_0002 esta en el diccionario.')

    # --------------------------------------------------------------------------------------------------------------

    # --------------------------------------------------------------------------------------------------------------
    # Si la variable ACTIVE_0001 contiene el valor TRUE se procede a eliminar todos los archivos de la carpeta 
    # especificada en la variable PATH_SALIDA_0001
    # --------------------------------------------------------------------------------------------------------------
    if continuar == True:
        if str(ACTIVE_0001).upper() == 'TRUE':

            logging.info(f'    Ha seteado la variable -ACTIVE_0002- como -TRUE- se eliminarán los archivos de la carpeta: {PATH_SALIDA_0001}.')
            
            conteo = 0
            for filename in os.listdir(PATH_SALIDA_0001):
                file_path = os.path.join(PATH_SALIDA_0001, filename)
                try:
                    os.remove(file_path)
                    logging.info(f'    Archivo {file_path} eliminado correctamente.')
                    conteo += 1
                except Exception as e:
                    logging.error(f'    Error al eliminar el archivo {file_path}: {e}')
                    conteo += 1

            if conteo == 0:
                logging.info(f'    No se han encontrado archivos en la carpeta {PATH_SALIDA_0001}')
        
    logging.warning('')
    logging.warning(' ** Finalizando LIMPIAR_OUTPUT_0001.. **')

# Fin de la funcion LIMPIAR_OUTPUT_0001
################################################################################################################


################################################################################################################
# Funcion de unificacion de tablas
def UNIFICADOR_TABLAS_0002(pasos_bot):

    logging.warning(' ** Ejecutando UNIFICADOR_TABLAS_0002.. **')
    logging.warning('')

    continuar = True

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de existencia de variables
    if continuar == True:

        ACTIVE_0002 = pasos_bot.get('ACTIVE_0002')
        if ACTIVE_0002 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave ACTIVE_0002 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave ACTIVE_0002 esta en el diccionario.')

        PATH_ENTRADA_0002 = pasos_bot.get('PATH_ENTRADA_0002')
        if PATH_ENTRADA_0002 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave PATH_ENTRADA_0002 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave PATH_ENTRADA_0002 esta en el diccionario.')

        PATH_SALIDA_0002 = pasos_bot.get('PATH_SALIDA_0002')
        if PATH_SALIDA_0002 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave PATH_SALIDA_0002 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave PATH_SALIDA_0002 esta en el diccionario.')

        NAME_0002 = pasos_bot.get('NAME_0002')
        if NAME_0002 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave NAME_0002 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave NAME_0002 esta en el diccionario.')

        EXPORT_0002 = pasos_bot.get('EXPORT_0002')
        if EXPORT_0002 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave EXPORT_0002 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave EXPORT_0002 esta en el diccionario.')
    # --------------------------------------------------------------------------------------------------------------

    # --------------------------------------------------------------------------------------------------------------
    # Funcion de recoleccion de archivos excels de una carpeta especifica.
    if str(ACTIVE_0002).upper() == "TRUE":

        logging.info('    Ejecutando Funcion UNIFICADOR_TABLAS...')
        logging.info(f'    PATH_ENTRADA_0002= {PATH_ENTRADA_0002}')
        logging.info(f'    PATH_SALIDA_0002= {PATH_SALIDA_0002}')
        logging.info(f'    EXPORT_0002= {EXPORT_0002}')
        logging.info(f'    ACTIVE_0002= {ACTIVE_0002}')
        logging.info(f'    NAME_0002= {NAME_0002}')

        # --------------------------------------------------------------------------------------------------------------
        # Verificar que exista EN_PATH_ENTRADA
        if continuar == True:
        
            if PATH_ENTRADA_0002:
                logging.info(f'    Recolectar todos los archivos .xlsx de la carpeta: {PATH_ENTRADA_0002}')
                
                ruta = str(PATH_ENTRADA_0002)
                if os.path.isdir(ruta):
                        logging.info(f'    La carpeta "{PATH_ENTRADA_0002}" existe.')
                else:
                    messagebox.showinfo('Error', f'La carpeta de entrada "{PATH_ENTRADA_0002}" no existe.')
                    continuar = False 

            else:
                logging.error(f'ERROR-    No se ha especificado la variable PATH_ENTRADA_0002 (str). PATH_ENTRADA_0002="C/PATH/FOLDER/INPUT"')
                messagebox.showinfo('Error', f'La carpeta de entrada "{PATH_ENTRADA_0002}" no existe.')
                continuar = False   

            logging.info(f'    verificando existencia de la carpeta: {PATH_SALIDA_0002}.')
            
            try:
                os.makedirs(PATH_SALIDA_0002, exist_ok=True)
            except Exception as e:
                messagebox.showerror('Error',f'Error al intentar crear la carpeta: {PATH_SALIDA_0002}. Detalles del error: {e}')
        # --------------------------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------------------------
        # En caso de que la variable EN_EXPORT=True (bool) se procede a validar el nombre del archivo de salida
        if continuar == True:

            if str(EXPORT_0002).upper =="TRUE":

                PATH_SALIDA_0002 = f'{PATH_SALIDA_0002}/{NAME_0002}'

                logging.info(f'    Se exportará el archivo unificado en la carpeta: {PATH_SALIDA_0002}')
            else:
                if str(EXPORT_0002).upper =="TRUE":
                    logging.info(f'    No se ha especificado nombre para el archivo. EXPORT_0002="Reporte.xlsx".')
        # --------------------------------------------------------------------------------------------------------------
            
        # --------------------------------------------------------------------------------------------------------------
        # Verificar que existan 1 o más excels en la ruta para iterar.   
        if continuar == True:

            # Obtener la lista de archivos Excel en la carpeta
            archivos_excel = [archivo for archivo in os.listdir(PATH_ENTRADA_0002) if archivo.endswith('.xlsx')]

            # Validar que existan excels a procesar
            if archivos_excel:
                logging.info(f'    Archivos a procesar: {archivos_excel}')
            else:
                logging.error(f'    No existen archivos a procesar en la carpeta {PATH_ENTRADA_0002}')
                continuar = False
        # --------------------------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------------------------
        # Bucle para recolectar los inputs en un solo DF   
        if continuar == True:
                
                logging.info(f'    Creando DataFrame en pandas...')

                # Crear un DataFrame vacío donde se unirán todos los datos de Excel
                df_completo = pd.DataFrame()

                # Recorrer todos los archivos Excel en la carpeta
                for archivo in archivos_excel:

                    logging.info(f'    Archivo a procesar: {archivo}')
                    try:
                        # Leer el archivo Excel en un DataFrame de pandas
                        df_excel = pd.read_excel(os.path.join(PATH_ENTRADA_0002, archivo))

                        # Unir los datos del archivo Excel al DataFrame completo
                        df_completo = pd.concat([df_completo, df_excel])

                        logging.info(f'    Procesado exitosamente: {archivo}')

                    except Exception as e:
                        logging.error(f'    No se pudo procesar el archivo: {archivo}')
                        print(f'    Error al leer el archivo "{archivo}". Detalle del error: {e}')

                logging.info(f'    Todos los archivos fueron procesados.')
        # --------------------------------------------------------------------------------------------------------------

        # --------------------------------------------------------------------------------------------------------------
        # Exportar Resultados   
        if continuar == True and str(EXPORT_0002).upper() == "TRUE":
        # Guardar el DataFrame completo en un nuevo archivo Excel
        
            try:
                PATH_SALIDA_0002 = f'{PATH_SALIDA_0002}/{NAME_0002}'
                df_completo.to_excel(PATH_SALIDA_0002, index=False)
                logging.info(f'    Todos los archivos .xlsx se han unido en "{PATH_SALIDA_0002}".')
            except Exception as e:
                logging.error(f'    No se pudo crear el archivo: {PATH_SALIDA_0002} - error: {e}')
        # --------------------------------------------------------------------------------------------------------------
            
        # --------------------------------------------------------------------------------------------------------------
        # Cerrar funcion 
        logging.warning('')
        logging.warning(' ** Finalizando UNIFICADOR_TABLAS_0002.. **') 

        return df_completo
         # --------------------------------------------------------------------------------------------------------------
    # --------------------------------------------------------------------------------------------------------------
    
# Fin de la funcion UNIFICADOR_TABLAS_0002
################################################################################################################


################################################################################################################
# Funcion de unificacion de tablas
def IMPORTAR_DATAFRAME_0003(pasos_bot):

    logging.warning(' ** Ejecutando IMPORTAR_DATAFRAME_0003.. **')
    logging.warning('')
    continuar = True

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de existencia de variables
    if continuar == True:

        ACTIVE_0003 = pasos_bot.get('ACTIVE_0003')
        if ACTIVE_0003 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave ACTIVE_0003 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave ACTIVE_0003 esta en el diccionario.')

    if str(ACTIVE_0003).upper() == "TRUE":

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de variables
        if continuar:
            PATH_ENTRADA_0003 = pasos_bot.get('PATH_ENTRADA_0003')
            if PATH_ENTRADA_0003 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave PATH_ENTRADA_0003 no está presente en el diccionario.')
                continuar = False

            PATH_SALIDA_0003 = pasos_bot.get('PATH_SALIDA_0003')
            if PATH_SALIDA_0003 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave PATH_SALIDA_0003 no está presente en el diccionario.')
                continuar = False

            HOJA_0003 = pasos_bot.get('HOJA_0003')
            if HOJA_0003 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave HOJA_0003 no está presente en el diccionario.')
                continuar = False

            EXPORT_0003 = pasos_bot.get('EXPORT_0003')
            if EXPORT_0003 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave EXPORT_0003 no está presente en el diccionario.')
                continuar = False

            TYPE_0003 = pasos_bot.get('TYPE_0003')
            if TYPE_0003 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave TYPE_0003 no está presente en el diccionario.')
                continuar = False
        # --------------------------------------------------------------------------------------------------------------

        if continuar:

            logging.info(f'')
            logging.info(f'    PATH_ENTRADA_0003= {PATH_ENTRADA_0003}')
            logging.info(f'    PATH_SALIDA_0003= {PATH_SALIDA_0003}')
            logging.info(f'    HOJA_0003= {HOJA_0003}')
            logging.info(f'    TYPE_0003= {TYPE_0003}')
            logging.info(f'    ACTIVE_0003= {ACTIVE_0003}')
            logging.info(f'    EXPORT_0003= {EXPORT_0003}')
            logging.info(f'')

            # --------------------------------------------------------------------------------------------------------------
            # Verificar que exista archivo entrada PATH_ENTRADA_0003
            if continuar == True:
                
                if os.path.isfile(PATH_ENTRADA_0003):
                    logging.info(f'    Importar DataFrame del archivo: {PATH_ENTRADA_0003}, Hoja: {HOJA_0003}')
                else:
                    messagebox.showinfo('Error | No lectura de input',f'No se ha encontrado un archivo en el path: {PATH_ENTRADA_0003}. Verifique la variable "PATH_ENTRADA_0003"')
                    logging.info('    Finalizando funcion..')
                    continuar = False
                    
                logging.info(f'    verificando existencia de la carpeta: {PATH_SALIDA_0003}.En caso que no exisitiese se procede a crearla.')

                os.makedirs(PATH_SALIDA_0003, exist_ok=True)
            # --------------------------------------------------------------------------------------------------------------

            # --------------------------------------------------------------------------------------------------------------
            # Crear DataFrame
            if continuar == True:
            
                logging.info(f'    Creando data frame en pandas.')

                try:
                    # Cargar el libro de trabajo (workbook) de Excel
                    workbook = load_workbook(filename=PATH_ENTRADA_0003, read_only=True)
                    
                    # Verificar si la hoja existe en el libro de trabajo
                    if HOJA_0003 in workbook.sheetnames:
                        logging.info(f'    La hoja {HOJA_0003} existe en el archivo {PATH_ENTRADA_0003}.')
                        
                        df_excel = pd.read_excel(PATH_ENTRADA_0003, sheet_name=HOJA_0003)
                
                    else:
                        logging.error(f'ERROR | IMPORTAR_DATAFRAME_0003 | La hoja {HOJA_0003} no existe en el archivo {PATH_ENTRADA_0003}.')
                        logging.error('    Finalizando función..')
                        continuar = False
                    
                except Exception as e:
                    logging.warning(f'No se pudo procesar el archivo: {PATH_ENTRADA_0003}')
                    print(f'Error al leer el archivo "{PATH_ENTRADA_0003}": {e}')
                    continuar = False

            # --------------------------------------------------------------------------------------------------------------

            # --------------------------------------------------------------------------------------------------------------
            # En caso de que la variable EN_EXPORT=True (bool) se procede a validar el nombre del archivo de salida
            if continuar == True:
                if str(EXPORT_0003).upper() == "TRUE":

                    PATH_SALIDA_0003 = f'{PATH_SALIDA_0003}/IMPORTAR_DATAFRAME_0003.xlsx'
                    logging.info(f'    Se exportará el archivo unificado en la carpeta: {PATH_SALIDA_0003}')

                    try:
                        df_excel.to_excel(PATH_SALIDA_0003, index=False)
                        logging.info(f'    Exportando... "{PATH_SALIDA_0003}".')
                    except Exception as e:
                        logging.error(f'    No se pudo crear el archivo: {PATH_SALIDA_0003} - error: {e}')
                
                logging.warning('')
                logging.warning(' ** Finalizando IMPORTAR_DATAFRAME_0003.. **')
                logging.warning('')

                return df_excel
            logging.warning('')
            logging.warning(' ** Finalizando IMPORTAR_DATAFRAME_0003.. **')
            logging.warning('')

    else:
        logging.info('    Para que la funcion "IMPORTAR_DATAFRAME_0003" se active debe colocar "TRUE" en la variable "ACTIVE_0003" del archivo de configuraciones.')
        logging.info('    Finalizando funcion...')

# Fin de la funcion
################################################################################################################


################################################################################################################
# Funcion de unificacion de tablas
def IMPORTAR_COLUMNA_0004(pasos_bot):

    logging.warning(' ** Ejecutando IMPORTAR_COLUMNA_0004.. **')
    logging.warning('')
    continuar = True

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de existencia de variables
    if continuar == True:

        ACTIVE_0004 = pasos_bot.get('ACTIVE_0004')
        if ACTIVE_0004 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave ACTIVE_0004 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave ACTIVE_0004 esta en el diccionario.')

    if str(ACTIVE_0004).upper() == "TRUE":

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de variables
        if continuar:
            PATH_ENTRADA_0004 = pasos_bot.get('PATH_ENTRADA_0004')
            if PATH_ENTRADA_0004 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave PATH_ENTRADA_0004 no está presente en el diccionario.')
                continuar = False

            FOLDER_OUTPUT_0004 = pasos_bot.get('FOLDER_OUTPUT_0004')
            if FOLDER_OUTPUT_0004 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave FOLDER_OUTPUT_0004 no está presente en el diccionario.')
                continuar = False

            COLUMN_IMPORT_0004 = pasos_bot.get('COLUMN_IMPORT_0004')
            if COLUMN_IMPORT_0004 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave COLUMN_IMPORT_0004 no está presente en el diccionario.')
                continuar = False

            EXPORT_0004 = pasos_bot.get('EXPORT_0004')
            if EXPORT_0004 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave EXPORT_0004 no está presente en el diccionario.')
                continuar = False

            NAME_0004 = pasos_bot.get('NAME_0004')
            if NAME_0004 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave NAME_0004 no está presente en el diccionario.')
                continuar = False

            DATAFRAME_INPUT_0004 = pasos_bot.get('DATAFRAME_INPUT_0004')
            if DATAFRAME_INPUT_0004 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave DATAFRAME_INPUT_0004 no está presente en el diccionario.')
                continuar = False

        # --------------------------------------------------------------------------------------------------------------
        
        if continuar:

            logging.info(f'')
            logging.info(f'    DATAFRAME_INPUT_0004= {DATAFRAME_INPUT_0004}')
            logging.info(f'    PATH_ENTRADA_0004= {PATH_ENTRADA_0004}')
            logging.info(f'    FOLDER_OUTPUT_0004= {FOLDER_OUTPUT_0004}')
            logging.info(f'    COLUMN_IMPORT_0004= {COLUMN_IMPORT_0004}')
            logging.info(f'    EXPORT_0004= {EXPORT_0004}')
            logging.info(f'    NAME_0004= {NAME_0004}')
   
            logging.info(f'')

            
            # --------------------------------------------------------------------------------------------------------------
            # Verificar que exista archivo entrada PATH_ENTRADA_0004
            if continuar == True:
                
                if os.path.isfile(PATH_ENTRADA_0004):
                    logging.info(f'    Importar DataFrame del archivo: {PATH_ENTRADA_0004}, Hoja: {NAME_0004}')

                    # verificar existencia de hoja
                    if verificar_existencia_hoja(PATH_ENTRADA_0004, NAME_0004):
                        logging.info(f'    La hoja {NAME_0004} Existe.')
                    else:
                        messagebox.showinfo('Error | No lectura de input',f'No existe la hoja {NAME_0004} en el archivo {PATH_ENTRADA_0004}.')
                        continuar = False
                else:
                    messagebox.showinfo('Error | No lectura de input',f'No se ha encontrado un archivo vàlido en el path: {PATH_ENTRADA_0004}. Verifique la variable "PATH_ENTRADA_0003"')
                    logging.info('    Finalizando funcion..')
                    continuar = False
                    
                logging.info(f'    verificando existencia de la carpeta: {FOLDER_OUTPUT_0004}.En caso que no exisitiese se procede a crearla.')

                os.makedirs(FOLDER_OUTPUT_0004, exist_ok=True)
            # --------------------------------------------------------------------------------------------------------------

            # --------------------------------------------------------------------------------------------------------------
            # Importar columna
            
            if continuar == True:
                try:
                    if str(DATAFRAME_INPUT_0004).upper() == "IMPORTAR_DATAFRAME_0003":
                        logging.info('    Se intenta obtener el DataFrame de entrada con la funcion: IMPORTAR_DATAFRAME_0003')
                        df_excel = IMPORTAR_DATAFRAME_0003(pasos_bot)
                    else:
                        df_excel = UNIFICADOR_TABLAS_0002(pasos_bot)
                        logging.info('    Se intenta obtener el DataFrame de entrada con la funcion: UNIFICADOR_TABLAS_0002')

                    logging.info('    DataFrame creado exitosamente.')

                except Exception as e:
                    messagebox.showinfo(f'No se pudo generar el DataFrame de entrada. Detalles del error: {e}')
                    continuar = False

            # --------------------------------------------------------------------------------------------------------------
            # Incorporar la columna del archivo adicional

            if continuar == True:

                logging.info(f'    Se Importará la columna "{COLUMN_IMPORT_0004}" de la hoja "{NAME_0004}" del archivo "{PATH_ENTRADA_0004}".')

                try:
                    
                    # Importar datos del excel al DataFrame
                    try:
                        logging.info(f'    Importando excel "{PATH_ENTRADA_0004}"')
                        DF_INPUT = pd.read_excel(PATH_ENTRADA_0004, sheet_name=NAME_0004)
                        logging.info(f'    Importado exitosamente.')
                    except Exception as e:
                        logging.error(f'    Error al importar archivo: "{PATH_ENTRADA_0004}". Detalles: {str(e)}') 
                        continuar = False

                    if continuar == True:

                        #Verificar existencia de la columna en dataframe
                        if COLUMN_IMPORT_0004 in df_excel.columns:

                            messagebox.showinfo('Error de existencia de columna',f'Se ha encontrado el header "{COLUMN_IMPORT_0004}" dentro del Dataframe. Para incorporar la columna debe tener un nombre diferente.')
                            continuar = False
                        else: 
                            logging.info(f'    La columna con encabezado "{COLUMN_IMPORT_0004}" existe en el DataFrame.')
                            
                            logging.warning('    Incorporando columna "{COLUMN_IMPORT_0004}"...')

                            try:

                                if COLUMN_IMPORT_0004 not in df_excel.columns:
                                    # Incorporar la columna COLUMN_IMPORT_0004 al final del DataFrame DF_INPUT
                                    columna_importada = pd.read_excel(PATH_ENTRADA_0004, sheet_name=NAME_0004, usecols=[COLUMN_IMPORT_0004])
                                    DF_COMPLETO = pd.concat([df_excel, columna_importada], axis=1)
                                else:
                                    messagebox.showinfo('Error al incorporar columna',f'Ya existe una columna llamada "{COLUMN_IMPORT_0004}" en el DataFrame destino.')
                                    continuar = False

                            except Exception as e:
                                messagebox.showinfo('Error al incorporar columna',f'La columna no pudo ser añidada al DataFrame. Detalles del error: {e}')
                                continuar = False



                except Exception as e:
                    messagebox.showinfo(f'No se pudo importar la columna. Detalles del error: {e}')
                    continuar = False

            # --------------------------------------------------------------------------------------------------------------
            # En caso de que la variable EN_EXPORT=True (bool) se procede a validar el nombre del archivo de salida
            if continuar == True:
                if str(EXPORT_0004).upper() == "TRUE":

                    FOLDER_OUTPUT_0004 = f'{FOLDER_OUTPUT_0004}/IMPORTAR_COLUMNA_0004.xlsx'
                    logging.info(f'    Se exportará el archivo unificado en la carpeta: {FOLDER_OUTPUT_0004}')

                    try:
                        DF_COMPLETO.to_excel(FOLDER_OUTPUT_0004, index=False)
                        logging.info(f'    Exportando... "{FOLDER_OUTPUT_0004}".')
                    except Exception as e:
                        logging.error(f'    No se pudo crear el archivo: {FOLDER_OUTPUT_0004} - error: {e}')
                
                logging.warning('')
                logging.warning(' ** Finalizando IMPORTAR_COLUMNA_0004.. **')
                logging.warning('')

                return DF_COMPLETO
            
            logging.warning('')
            logging.warning(' ** Finalizando IMPORTAR_COLUMNA_0004.. **')
            logging.warning('')

    else:
        logging.info('    Para que la funcion "IMPORTAR_COLUMNA_0004" se active debe colocar "TRUE" en la variable "ACTIVE_0004" del archivo de configuraciones.')
        logging.info('    Finalizando funcion...')

# Fin de la funcion
################################################################################################################


################################################################################################################
# Funcion de unificacion de tablas
def IMPORTAR_COLUMNA_CRUCE_0005(pasos_bot):

    logging.warning(' ** Ejecutando IMPORTAR_COLUMNA_CRUCE_0005.. **')
    logging.warning('')
    continuar = True

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de existencia de variables
    if continuar == True:

        ACTIVE_0005 = pasos_bot.get('ACTIVE_0005')
        if ACTIVE_0005 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave ACTIVE_0005 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave ACTIVE_0005 esta en el diccionario.')

    if str(ACTIVE_0005).upper() == "TRUE":

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de variables
        if continuar:
            DATAFRAME_INPUT_0005 = pasos_bot.get('DATAFRAME_INPUT_0005')
            if DATAFRAME_INPUT_0005 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave DATAFRAME_INPUT_0005 no está presente en el diccionario.')
                continuar = False


            PATH_ENTRADA_0005 = pasos_bot.get('PATH_ENTRADA_0005')
            if PATH_ENTRADA_0005 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave PATH_ENTRADA_0005 no está presente en el diccionario.')
                continuar = False

            FOLDER_OUTPUT_0005 = pasos_bot.get('FOLDER_OUTPUT_0005')
            if FOLDER_OUTPUT_0005 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave FOLDER_OUTPUT_0005 no está presente en el diccionario.')
                continuar = False

            HEADER_DF_COMPARE_0005 = pasos_bot.get('HEADER_DF_COMPARE_0005')
            if HEADER_DF_COMPARE_0005 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave HEADER_DF_COMPARE_0005 no está presente en el diccionario.')
                continuar = False

            HEADER_IN_COMPARE_0005 = pasos_bot.get('HEADER_IN_COMPARE_0005')
            if HEADER_IN_COMPARE_0005 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave HEADER_IN_COMPARE_0005 no está presente en el diccionario.')
                continuar = False

            HEADER_IMPORT_0005 = pasos_bot.get('HEADER_IMPORT_0005')
            if HEADER_IMPORT_0005 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave HEADER_IMPORT_0005 no está presente en el diccionario.')
                continuar = False

            NAME_0005 = pasos_bot.get('NAME_0005')
            if NAME_0005 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave NAME_0005 no está presente en el diccionario.')
                continuar = False

            EXPORT_0005 = pasos_bot.get('EXPORT_0005')
            if EXPORT_0005 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave EXPORT_0005 no está presente en el diccionario.')
                continuar = False


        # --------------------------------------------------------------------------------------------------------------
        
        if continuar:

            logging.info(f'')
            logging.info(f'    DATAFRAME_INPUT_0005= {DATAFRAME_INPUT_0005}')
            logging.info(f'    PATH_ENTRADA_0005= {PATH_ENTRADA_0005}')
            logging.info(f'    FOLDER_OUTPUT_0005= {FOLDER_OUTPUT_0005}')
            logging.info(f'    HEADER_DF_COMPARE_0005= {HEADER_DF_COMPARE_0005}')
            logging.info(f'    HEADER_IN_COMPARE_0005= {HEADER_IN_COMPARE_0005}')
            logging.info(f'    HEADER_IMPORT_0005= {HEADER_IMPORT_0005}')
            logging.info(f'    NAME_0005= {NAME_0005}')
   
            logging.info(f'')

            
            # --------------------------------------------------------------------------------------------------------------
            # Verificar que exista archivo entrada PATH_ENTRADA_0005
            if continuar == True:
                
                if os.path.isfile(PATH_ENTRADA_0005):
                    logging.info(f'    Importar DataFrame del archivo: {PATH_ENTRADA_0005}, Hoja: {NAME_0005}')

                    # verificar existencia de hoja
                    if verificar_existencia_hoja(PATH_ENTRADA_0005, NAME_0005):
                        logging.info(f'    La hoja {NAME_0005} Existe.')
                    else:
                        messagebox.showinfo('Error | No lectura de input',f'No existe la hoja {NAME_0005} en el archivo {PATH_ENTRADA_0005}.')
                        continuar = False
                else:
                    messagebox.showinfo('Error | No lectura de input',f'No se ha encontrado un archivo vàlido en el path: {PATH_ENTRADA_0005}. Verifique la variable "PATH_ENTRADA_0003"')
                    logging.info('    Finalizando funcion..')
                    continuar = False
                    
                logging.info(f'    verificando existencia de la carpeta: {FOLDER_OUTPUT_0005}.En caso que no exisitiese se procede a crearla.')

                os.makedirs(FOLDER_OUTPUT_0005, exist_ok=True)
            # --------------------------------------------------------------------------------------------------------------

            
            # --------------------------------------------------------------------------------------------------------------
            # Verificar que exista columna a comparar en DF input
            if continuar == True:

                try:
                    DataFrameInput = pd.read_excel(DATAFRAME_INPUT_0005)
                except Exception as e:
                    logging.error(f'    Error en copiar DATAFRAME_INPUT_0005. Error: {e}')
                
                if HEADER_DF_COMPARE_0005 not in DataFrameInput.columns:
                    logging.info(f'    La funcion no puede continuar debido a que el header "{HEADER_DF_COMPARE_0005}" no existe en el Data frame de entrada 1.')
                    return HEADER_DF_COMPARE_0005
                else: 
                    logging.info(f'La columna con encabezado "{HEADER_DF_COMPARE_0005}" existe en el DataFrame.')

            # --------------------------------------------------------------------------------------------------------------

            
            # --------------------------------------------------------------------------------------------------------------
            # Verificar que existan columnas a incoprporar en DF cruce
            if continuar == True:

                try:
                    logging.info('    importando archivo: {PATH_ENTRADA_0005}')
                    DataFrameInput2 = pd.read_excel(PATH_ENTRADA_0005,usecols=[HEADER_IN_COMPARE_0005,HEADER_IMPORT_0005])
                except Exception as e:
                    logging.error('    Error en importar  PATH_ENTRADA_0005. Error: {e}')
                
                if HEADER_IN_COMPARE_0005 not in DataFrameInput2.columns:
                    if HEADER_IMPORT_0005 not in DataFrameInput2.columns:
                        logging.info(f'    La funcion no puede continuar debido a que el header "{HEADER_IMPORT_0005}" no existe en el Data frame de entrada 1.')
                    else:
                        logging.info(f'    La funcion no puede continuar debido a que el header "{HEADER_IN_COMPARE_0005}" no existe en el Data frame de entrada 1.')
                    continuar = False
                else: 
                    logging.info(f'    Las columnas "{HEADER_DF_COMPARE_0005}" y "{HEADER_IMPORT_0005}" existe en el DataFrame.')

            # --------------------------------------------------------------------------------------------------------------

            newDataFrame = DataFrameInput.copy()

            if continuar == True:
                for index, row in DataFrameInput.iterrows():
                    
                    
                    value1 = row[HEADER_DF_COMPARE_0005]

                    value2 = DataFrameInput2.loc[DataFrameInput2[HEADER_IN_COMPARE_0005] == value1, HEADER_IMPORT_0005].values
                    if len(value2) > 0:
                        try:
                            newDataFrame.at[index, HEADER_IMPORT_0005] = value2[0]
                        except Exception as e:
                            logging('    error : {e}')

            # --------------------------------------------------------------------------------------------------------------
            # En caso de que la variable EN_EXPORT=True (bool) se procede a validar el nombre del archivo de salida
            if continuar == True:
                if str(EXPORT_0005).upper() == "TRUE":

                    FOLDER_OUTPUT_0005 = f'{FOLDER_OUTPUT_0005}/IMPORTAR_COLUMNA_CRUCE_0005.xlsx'
                    logging.info(f'    Se exportará el archivo unificado en la carpeta: {FOLDER_OUTPUT_0005}')

                    try:
                        newDataFrame.to_excel(FOLDER_OUTPUT_0005, index=False)
                        logging.info(f'    Exportando... "{FOLDER_OUTPUT_0005}".')
                    except Exception as e:
                        logging.error(f'    No se pudo crear el archivo: {FOLDER_OUTPUT_0005} - error: {e}')
                
                logging.warning('')
                logging.warning(' ** Finalizando IMPORTAR_COLUMNA_CRUCE_0005.. **')
                logging.warning('')

                return newDataFrame
            
            logging.warning('')
            logging.warning(' ** Finalizando IMPORTAR_COLUMNA_CRUCE_0005.. **')
            logging.warning('')

    else:
        logging.info('    Para que la funcion "IMPORTAR_COLUMNA_CRUCE_0005" se active debe colocar "TRUE" en la variable "ACTIVE_0005" del archivo de configuraciones.')
        logging.info('    Finalizando funcion...')

# Fin de la funcion
################################################################################################################


################################################################################################################
# Funcion de Validaciones generales
def VALIDACIONES_GENERALES_0006(pasos_bot):

    logging.warning(' ** Ejecutando VALIDACIONES_GENERALES_0006.. **')
    logging.warning('')
    continuar = True

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de existencia de variables
    if continuar == True:

        ACTIVE_0006 = pasos_bot.get('ACTIVE_0006')
        if ACTIVE_0006 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave ACTIVE_0005 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave ACTIVE_0006 esta en el diccionario.')

    if str(ACTIVE_0006).upper() == "TRUE":

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de variables
        if continuar:

            PATH_ENTRADA_0006 = pasos_bot.get('PATH_ENTRADA_0006')
            if PATH_ENTRADA_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave PATH_ENTRADA_0006 no está presente en el diccionario.')
                continuar = False

            FOLDER_OUTPUT_0006 = pasos_bot.get('FOLDER_OUTPUT_0006')
            if FOLDER_OUTPUT_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave FOLDER_OUTPUT_0006 no está presente en el diccionario.')
                continuar = False

            HEADER_TI_0006 = pasos_bot.get('HEADER_TI_0006')
            if HEADER_TI_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave HEADER_TI_0006 no está presente en el diccionario.')
                continuar = False

            HEADER_SOC_0006 = pasos_bot.get('HEADER_SOC_0006')
            if HEADER_SOC_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave HEADER_SOC_0006 no está presente en el diccionario.')
                continuar = False

            NAME_0006 = pasos_bot.get('NAME_0006')
            if NAME_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave NAME_0006 no está presente en el diccionario.')
                continuar = False

            EXPORT_0006 = pasos_bot.get('EXPORT_0006')
            if EXPORT_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave EXPORT_0006 no está presente en el diccionario.')
                continuar = False

            TABLA_INVERSIONES_0006 = pasos_bot.get('TABLA_INVERSIONES_0006')
            if TABLA_INVERSIONES_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave TABLA_INVERSIONES_0006 no está presente en el diccionario.')
                continuar = False

            HEADER_COMPARATE_0006 = pasos_bot.get('HEADER_COMPARATE_0006')
            if HEADER_COMPARATE_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave HEADER_COMPARATE_0006 no está presente en el diccionario.')
                continuar = False


            VACIO_01_0006 = pasos_bot.get('VACIO_01_0006')
            if VACIO_01_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave VACIO_01_0006 no está presente en el diccionario.')
                continuar = False

            VACIO_02_0006 = pasos_bot.get('VACIO_02_0006')
            if VACIO_02_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave VACIO_02_0006 no está presente en el diccionario.')
                continuar = False


            DIFERENTE_01_0006 = pasos_bot.get('DIFERENTE_01_0006')
            if DIFERENTE_01_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave DIFERENTE_01_0006 no está presente en el diccionario.')
                continuar = False

            PARAMETRO_DIF_01_0006 = pasos_bot.get('PARAMETRO_DIF_01_0006')
            if PARAMETRO_DIF_01_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave PARAMETRO_DIF_01_0006 no está presente en el diccionario.')
                continuar = False

            FIRST_2_01_0006 = pasos_bot.get('FIRST_2_01_0006')
            if FIRST_2_01_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave FIRST_2_01_0006 no está presente en el diccionario.')
                continuar = False

            FIRST_2_02_0006 = pasos_bot.get('FIRST_2_02_0006')
            if FIRST_2_02_0006 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave FIRST_2_02_0006 no está presente en el diccionario.')
                continuar = False

        # --------------------------------------------------------------------------------------------------------------
        
        if continuar:

            logging.info(f'')

            logging.info(f'    PATH_ENTRADA_0006= {PATH_ENTRADA_0006}')
            logging.info(f'    NAME_0006= {NAME_0006}')
            logging.info(f'    FOLDER_OUTPUT_0006= {FOLDER_OUTPUT_0006}')

            logging.info(f'')

            # --------------------------------------------------------------------------------------------------------------
            # Verificar que exista archivo entrada PATH_ENTRADA_0006
            if continuar == True:
                
                if os.path.isfile(PATH_ENTRADA_0006):
                    logging.info(f'    Importar DataFrame del archivo: {PATH_ENTRADA_0006}, Hoja: {NAME_0006}')

                    # verificar existencia de hoja
                    if verificar_existencia_hoja(PATH_ENTRADA_0006, NAME_0006):
                        logging.info(f'    La hoja {NAME_0006} Existe.')
                    else:
                        messagebox.showinfo('Error | No lectura de input',f'No existe la hoja {NAME_0006} en el archivo {PATH_ENTRADA_0006}.')
                        continuar = False
                else:
                    messagebox.showinfo('Error | No lectura de input',f'No se ha encontrado un archivo vàlido en el path: {PATH_ENTRADA_0006}. Verifique la variable "PATH_ENTRADA_0003"')
                    logging.info('    Finalizando funcion..')
                    continuar = False
                    
                logging.info(f'    verificando existencia de la carpeta: {FOLDER_OUTPUT_0006}.En caso que no exisitiese se procede a crearla.')

                os.makedirs(FOLDER_OUTPUT_0006, exist_ok=True)
            # --------------------------------------------------------------------------------------------------------------

            # --------------------------------------------------------------------------------------------------------------
            # Verificar que exista columna a comparar en DF input
            if continuar == True:

                try:
                    DataFrameInput = pd.read_excel(PATH_ENTRADA_0006,sheet_name=NAME_0006)
                except Exception as e:
                    logging.error(f'    Error en copiar DATAFRAME_INPUT_0005. Error: {e}')

            # --------------------------------------------------------------------------------------------------------------

            
            # --------------------------------------------------------------------------------------------------------------
            # Leyendo tabla de inversiones
            if continuar == True:
                
                print(f'Leyendo tabla de Inversiones. Path: {TABLA_INVERSIONES_0006}')

                try:
                    #Diccionario donde se guardara la tabla de inversiones
                    Tabla_Inversiones={}

                    archivo_excel = pd.ExcelFile(TABLA_INVERSIONES_0006)
                    df = archivo_excel.parse('Tabla_Inversiones')
                    print(f'Leyendo Variables "INVERSION" y "LETRA"')
                    for index, row in df.iterrows():
                        key = row['INVERSION']
                        value = row['LETRA']
                        Tabla_Inversiones[key] = value
                    
                    continuar = True
                    print(f'Lectura de Config finalizada correctamente.')

                except Exception as e:
                    continuar = False
                    print(f'Hubo un error al leer el Config: {e}')

            # --------------------------------------------------------------------------------------------------------------

            # --------------------------------------------------------------------------------------------------------------
            # Validaciones generales
            if continuar == True:
                
                logging.info(f'    En caso que el DataFrameInput no contenga la columna "Comentarios", se procede a crearla al final..')
                
                # Verificar si la columna "Comentarios" no existe en el DataFrameInput
                if not any(column == 'Comentarios' for column in DataFrameInput.columns):
                # Crear la columna "Comentarios" con valor inicial vacío si no existe
                    DataFrameInput = DataFrameInput.assign(Comentarios='')


                VACIO_1 = pasos_bot.get('VACIO_01_0006')
                VACIO_2 = pasos_bot.get('VACIO_02_0006')

                DIFERENTE_1 = pasos_bot.get('DIFERENTE_01_0006')
                DIFERENTE_PARAMETRO_1 = pasos_bot.get('PARAMETRO_DIF_01_0006')

                F2_COL_1 =  pasos_bot.get('FIRST_2_01_0006')
                F2_COL_2 =  pasos_bot.get('FIRST_2_02_0006')

                for index, row in DataFrameInput.iterrows():
                    
                    # Validar que la columna VACIO_1 o VACIO_2 contenga datos.
                    if pd.isnull(row[VACIO_1]) or row[VACIO_1] == '':
                        DataFrameInput.at[index, 'Comentarios'] += f'| Valor esperado en Columna: "{VACIO_1}" |'

                    if pd.isnull(row[VACIO_2]) or row[VACIO_2] == '':
                        DataFrameInput.at[index, 'Comentarios'] += f'| Valor esperado en Columna: "{VACIO_2}" |'          

                    # Validar que la columna DIFERENTE_1 contenga el valor DIFERENTE_PARAMETRO_1.
                    if row[DIFERENTE_1] != DIFERENTE_PARAMETRO_1:
                        DataFrameInput.at[index, 'Comentarios'] += f'| Valor esperado en Columna "{DIFERENTE_1}" : "{DIFERENTE_PARAMETRO_1}" |'

                    # Validar que los primeros dos digitos de la columna F2_COL_1 coincidan con 
                    # los primeros dos digitos de la columna F2_COL_2.
                    if str(row[F2_COL_1])[:2] != str(row[F2_COL_2])[:2]:
                        if str(row[F2_COL_2]) == "" or pd.isnull(row[F2_COL_2]):
                            DataFrameInput.at[index, 'Comentarios'] += f'| No se encontró valor en la columna: {F2_COL_2}" |'
                        else:
                            if str(row[F2_COL_2]).upper() != "NO":
                                DataFrameInput.at[index, 'Comentarios'] += f'| El CECO debe comenzar con: {str(row[F2_COL_2])[:2]}" |'

                    # Validar que la OIE termine con la LETRA del diccionario y finalice con los ultimos
                    # dos digitos de la columna FIN_COL_1

                    #Buscar Letra                                    
                    largo = len(str(row[HEADER_COMPARATE_0006]))


                    if largo > 0:

                        letra_OIE = str(row[HEADER_COMPARATE_0006])[-3:-2]
                        tipo_Inversion = str(row[HEADER_TI_0006])

                        #buscarlo en diccionario tablaInverison
                        try:
                            tipo_inversion = Tabla_Inversiones.get(tipo_Inversion)
                        except:
                            Letra = 'No encontrada'

                    if tipo_inversion:
                        Letra = tipo_inversion
                    else:
                        Letra = 'No encontrada'

                    if str(Letra) != str(letra_OIE):

                        if str(tipo_Inversion).upper() == "NONE":
                            pass
                        else:
                            DataFrameInput.at[index, 'Comentarios'] += f' | Letra OIE según tabla: "{Letra}", no coincidente con OIE actual: {letra_OIE} |'
                    
                    first2_SOCIEDAD = str(row[HEADER_COMPARATE_0006])[-2:]
                    comparateSoc = str(row[HEADER_SOC_0006])[:2]

                    if str(first2_SOCIEDAD) != str(comparateSoc):
                            #Si trae NO por defecto (cruce con Sociedades)
                            if str(comparateSoc).upper() == "NO":
                                pass
                            else:
                                DataFrameInput.at[index, 'Comentarios'] += f' | Ultimos 2 digitos de OIE esperados: "{comparateSoc}". |'


            # --------------------------------------------------------------------------------------------------------------
            # En caso de que la variable EN_EXPORT=True (bool) se procede a validar el nombre del archivo de salida
            if continuar == True:
                if str(EXPORT_0006).upper() == "TRUE":

                    FOLDER_OUTPUT_0006 = f'{FOLDER_OUTPUT_0006}/VALIDACIONES_GENERALES_0006.xlsx'
                    logging.info(f'    Se exportará el archivo unificado en la carpeta: {FOLDER_OUTPUT_0006}')

                    try:
                        DataFrameInput.to_excel(FOLDER_OUTPUT_0006, index=False)
                        logging.info(f'    Exportando... "{FOLDER_OUTPUT_0006}".')
                    except Exception as e:
                        logging.error(f'    No se pudo crear el archivo: {FOLDER_OUTPUT_0006} - error: {e}')
                
                logging.warning('')
                logging.warning(' ** Finalizando VALIDACIONES_GENERALES_0006.. **')
                logging.warning('')

                return DataFrameInput
            

            logging.warning('')
            logging.warning(' ** Finalizando VALIDACIONES_GENERALES_0006.. **')
            logging.warning('')

    else:
        logging.info('    Para que la funcion "VALIDACIONES_GENERALES_0006" se active debe colocar "TRUE" en la variable "ACTIVE_0006" del archivo de configuraciones.')
        logging.info('    Finalizando funcion...')

# Fin de la funcion
################################################################################################################

################################################################################################################
# Funcion de generacion de mascara
def GENERACION_MASCARA_0007(pasos_bot):

    logging.warning(' ** Ejecutando GENERACION_MASCARA_0007.. **')
    logging.warning('')
    continuar = True

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de existencia de variables
    if continuar == True:

        ACTIVE_0007 = pasos_bot.get('ACTIVE_0007')
        if ACTIVE_0007 is None:
            messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave ACTIVE_0005 no está presente en el diccionario.')
            continuar = False
        else:
            logging.info('    La clave ACTIVE_0007 esta en el diccionario.')

    if str(ACTIVE_0007).upper() == "TRUE":

    # --------------------------------------------------------------------------------------------------------------
    # Validaciones de variables
        if continuar:

            PATH_ENTRADA_0007 = pasos_bot.get('PATH_ENTRADA_0007')
            if PATH_ENTRADA_0007 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave PATH_ENTRADA_0007 no está presente en el diccionario.')
                continuar = False

            NAME_0007 = pasos_bot.get('NAME_0007')
            if NAME_0007 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave NAME_0007 no está presente en el diccionario.')
                continuar = False

            FOLDER_OUTPUT_0007 = pasos_bot.get('FOLDER_OUTPUT_0007')
            if FOLDER_OUTPUT_0007 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave FOLDER_OUTPUT_0007 no está presente en el diccionario.')
                continuar = False

            EXPORT_0007 = pasos_bot.get('EXPORT_0007')
            if EXPORT_0007 is None:
                messagebox.showinfo('Clave no encontrada en archivo de configuraciones','La clave EXPORT_0007 no está presente en el diccionario.')
                continuar = False          

        # --------------------------------------------------------------------------------------------------------------
        
        if continuar:

            logging.info(f'')

            logging.info(f'    FOLDER_OUTPUT_0007= {FOLDER_OUTPUT_0007}')

            logging.info(f'')

            # --------------------------------------------------------------------------------------------------------------
            # Verificar que exista archivo entrada PATH_ENTRADA_0006
            if continuar == True:
                
                if os.path.isfile(PATH_ENTRADA_0007):
                    logging.info(f'    Importar DataFrame del archivo: {PATH_ENTRADA_0007}, Hoja: {NAME_0007}')

                    # verificar existencia de hoja
                    if verificar_existencia_hoja(PATH_ENTRADA_0007, NAME_0007):
                        logging.info(f'    La hoja {PATH_ENTRADA_0007} Existe.')
                    else:
                        messagebox.showinfo('Error | No lectura de input',f'No existe la hoja {NAME_0007} en el archivo {PATH_ENTRADA_0007}.')
                        continuar = False
                else:
                    messagebox.showinfo('Error | No lectura de input',f'No se ha encontrado un archivo vàlido en el path: {PATH_ENTRADA_0007}. Verifique la variable "PATH_ENTRADA_0003"')
                    logging.info('    Finalizando funcion..')
                    continuar = False
                    
                logging.info(f'    verificando existencia de la carpeta: {FOLDER_OUTPUT_0007}.En caso que no exisitiese se procede a crearla.')

                os.makedirs(FOLDER_OUTPUT_0007, exist_ok=True)
            # --------------------------------------------------------------------------------------------------------------

            # --------------------------------------------------------------------------------------------------------------
            # Verificar que exista columna a comparar en DF input
            if continuar == True:

                try:
                    DataFrameInput = pd.read_excel(PATH_ENTRADA_0007,sheet_name=NAME_0007)
                except Exception as e:
                    logging.error(f'    Error al crear el DataFrame. Error: {e}')

            # --------------------------------------------------------------------------------------------------------------

            # --------------------------------------------------------------------------------------------------------------
            # Validaciones generales
            if continuar == True:
                
                logging.info(f'Creando máscara...')

                # Leer el archivo Excel y crear el DataFrameInput
                DataFrameInput = pd.read_excel(PATH_ENTRADA_0007, sheet_name=NAME_0007)

                comentarios_vacios = DataFrameInput['Comentarios'].isnull().all()

                if comentarios_vacios:
                    # Realizar la sumatoria de la columna "Horas" agrupando por las columnas especificadas
                    df_sumatoria = DataFrameInput.groupby(["Orden de compra","Posición", "OIE (Inversión)", "CECO"])["Horas"].sum().reset_index()

                    # Crear el nuevo DataFrame con la estructura deseada
                    df_export = pd.DataFrame(columns=["Orden de compra","Posición", "OIE (Inversión)", "Horas","NO ESTA", "CECO"])
                    df_export.loc[1] = ["Pedido", "Pos Ped","OIE", "Cantidad", "Porcentaje %","CECO"]
            
                    df_export = pd.concat([df_export, df_sumatoria], ignore_index=True)

                    # Reordenar las columnas para que "Horas" esté antes de "CECO"
                    df_export = df_export[["Orden de compra","Posición", "OIE (Inversión)", "Horas", "NO ESTA","CECO"]]

                else:
                    # Mostrar mensaje de error de comentarios sin resolver
                    messagebox.showinfo('Error al generar máscara','La máscara contiene comentarios sin resolver.')
                    # Puedes mostrar un MsgBox con el mensaje de error aquí

            #--------------------------------------------------------------------------------------------
            # En caso de que la variable EN_EXPORT=True (bool) se procede a validar el nombre del archivo de salida
            if continuar == True:
                if str(EXPORT_0007).upper() == "TRUE":

                    FOLDER_OUTPUT_0007 = f'{FOLDER_OUTPUT_0007}/Mascara.xlsx'
                    logging.info(f'    Se exportará el archivo unificado en la carpeta: {FOLDER_OUTPUT_0007}')

                    try:
                        df_export.to_csv(FOLDER_OUTPUT_0007, index=False)
                        logging.info(f'    Exportando... "{FOLDER_OUTPUT_0007}".')
                    except Exception as e:
                        logging.error(f'    No se pudo crear el archivo: {FOLDER_OUTPUT_0007} - error: {e}')
                
                logging.warning('')
                logging.warning(' ** Finalizando GENERACION_MASCARA_0007.. **')
                logging.warning('')

                return df_sumatoria
            

            logging.warning('')
            logging.warning(' ** Finalizando GENERACION_MASCARA_0007.. **')
            logging.warning('')

    else:
        logging.info('    Para que la funcion "GENERACION_MASCARA_0007" se active debe colocar "TRUE" en la variable "ACTIVE_0007" del archivo de configuraciones.')
        logging.info('    Finalizando funcion...')

# Fin de la funcion
################################################################################################################
# --------------------------------------------------------------------------------------------------------------
